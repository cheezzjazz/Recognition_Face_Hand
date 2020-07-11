import sys
sys.path.append('../insightface/deploy')
sys.path.append('../insightface/src/common')

from keras.models import load_model
from mtcnn.mtcnn import MTCNN
from imutils import paths
from openpyxl import Workbook
from src.hand_tracker import HandTracker
import face_preprocess
import numpy as np
import face_model
import argparse
import pickle
import time
import dlib
import cv2
import os

write_wb = Workbook()
write_ws = write_wb.active
write_ws.append(["Menu", "Name", "Score", "renew?"])
WINDOW = "Hand Tracking"
PALM_MODEL_PATH = "models/palm_detection_without_custom_op.tflite"
LANDMARK_MODEL_PATH = "models/hand_landmark.tflite"
ANCHORS_PATH = "models/anchors.csv"

POINT_COLOR = (0, 255, 0)
KEY_COLOR =(0,0,255)
CONNECTION_COLOR = (255, 0, 0)
NEWFACE_COLOR = (255, 255, 255)
FACEBOX_COLOR=(255,0,0)
THICKNESS = 2

ap = argparse.ArgumentParser()

ap.add_argument("--mymodel", default="outputs/my_model.h5",
    help="Path to recognizer model")
ap.add_argument("--le", default="outputs/le.pickle",
    help="Path to label encoder")
ap.add_argument("--embeddings", default="outputs/embeddings.pickle",
    help='Path to embeddings')
ap.add_argument("--video-out", default="../datasets/videos_output/stream_test.mp4",
    help='Path to output video')


ap.add_argument('--image-size', default='112,112', help='')
ap.add_argument('--model', default='../insightface/models/model-y1-test2/model,0', help='path to load model.')
ap.add_argument('--ga-model', default='', help='path to load model.')
ap.add_argument('--gpu', default=0, type=int, help='gpu id')
ap.add_argument('--det', default=1, type=int, help='mtcnn option, 1 means using R+O, 0 means detect from begining')
ap.add_argument('--flip', default=0, type=int, help='whether do lr flip aug')
ap.add_argument('--threshold', default=1.24, type=float, help='ver dist threshold')

args = ap.parse_args()

# Load embeddings and labels
data = pickle.loads(open(args.embeddings, "rb").read())

embeddings = np.array(data['embeddings'])
labels = np.array(data['names'])        	# 200625

indexNumber = 0

# Initialize detector
detector = MTCNN()

# Initialize faces embedding model
embedding_model = face_model.FaceModel(args)

# Load the classifier model
model = load_model('outputs/my_model.h5')

def findCosineDistance(vector1, vector2):
    """
    Calculate cosine distance between two vector
    """
    vec1 = vector1.flatten()
    vec2 = vector2.flatten()


    a = np.dot(vec1, vec2)    #200626   a = np.dot(vec1.T, vec2)
    b = np.dot(vec1.T, vec1)
    c = np.dot(vec2.T, vec2)
    return 1 - (a/(np.sqrt(b)*np.sqrt(c)))

def CosineSimilarity(test_vec, source_vecs):
    """
    Verify the similarity of one vector to group vectors of one class
    """
    cos_dist = 0
    curr = 0
    minimum = 10
    temp = -1
    index = 0
    for source_vec in source_vecs:
        # cos_dist += findCosineDistance(test_vec, source_vec)
        curr = findCosineDistance(test_vec, source_vec)
        if curr < minimum :
            temp = index
            minimum = curr
        cos_dist += curr
        index += 1
    return cos_dist/len(source_vecs), temp

def getMaxfacebox(bboxes):
    max_area = 0
    for bboxe in bboxes:
        bbox = bboxe["box"]
        bbox = np.array([bbox[0], bbox[1], bbox[0] + bbox[2], bbox[1] + bbox[3]])
        keypoints = bboxe["keypoints"]
        area = (bbox[2] - bbox[0]) * (bbox[3] - bbox[1])
        if area > max_area:
            max_bbox = bbox
            landmarks = keypoints
            max_area = area
    return max_bbox[0:4], landmarks, max_area

# Program to find most frequent element in a list 
def most_frequent(List):
    counter = 0
    num = List[0]
    for i in List:
        curr_frequency = List.count(i)
        if(curr_frequency > counter):
            counter = curr_frequency
            num = i
    return num

# Initialize some useful arguments
cosine_threshold = 1.45
proba_threshold = 0.91
dis_threshold = 0.4
comparing_num = 5
trackers = []
texts = []
trackers_In = []
texts_In = []
frames = 0

draw_points = None
handScore = 0

# Start streaming and recording
rtsp = 'rtsp://admin:**graphics@163.152.162.189:554/cam/realmonitor?channel1&subtype=1'
cap = cv2.VideoCapture(1)
capIn = cv2.VideoCapture(0) #input camera
# ipcap = cv2.VideoCapture(0)
# cv2.SetCaptureProperty(ipcap, CV_CAP_PROP_BUFFERSIZE, 5)
# cap.set(cv2.cv.CV_CAP_PROP_FPS, 5)

frame_width = int(cap.get(3))
frame_height = int(cap.get(4))
save_width = frame_width #600
save_height = int(save_width/frame_width*frame_height) #int(600/frame_width*frame_height)
#video_out = cv2.VideoWriter(args.video_out, cv2.VideoWriter_fourcc('M','J','P','G'), 1, (save_width,save_height))

#hand connection
connections = [
    (0, 1), (1, 2), (2, 3), (3, 4),
    (5, 6), (6, 7), (7, 8),
    (9, 10), (10, 11), (11, 12),
    (13, 14), (14, 15), (15, 16), 
    (17, 18), (18, 19), (19, 20),
    (0, 5), (5, 9), (9, 13), (13, 17), (0, 17)
]

#hand detector
hand_detector = HandTracker(
    PALM_MODEL_PATH,
    LANDMARK_MODEL_PATH,
    ANCHORS_PATH,
    box_shift=0.2,
    box_enlarge=1.3
)

previous_embedding = []
new_labels = []
new_embeddings = []
max_bbox_In = np.zeros(4)
max_bbox_Out = np.zeros(4)
prev_name = None
name = None
score_number = [] # all of hand scores
score_number2 = [] # 3 frames continue scores
when = 0
while True:
    # ret is fading
    ret, frame = cap.read()
    _, frame_In = capIn.read()
    #ret, ipframe = ipcap.read()
    frames += 1
    
    if frames % 5 == 0:
        rgb_In = cv2.cvtColor(frame_In, cv2.COLOR_BGR2RGB) #input camera
        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        #frame = cv2.resize(frame, (save_width, save_height))
        # ipframe = cv2.resize(ipframe, (640, 480))

        cnt = []
        # name = None

        detect_tick = time.time()
        bboxes_In = detector.detect_faces(frame_In) #input camera
        bboxes_Out = detector.detect_faces(frame)
        detect_tock = time.time()

        detect_tick = time.time()
        bboxes_Out = detector.detect_faces(frame) #output camera
        detect_tock = time.time()
        #ip_detect_tick = time.time()
        # ipbboxes = detector.detect_faces(ipframe)
        #ip_detect_tock = time.time()
        trackers_In = []
        texts_In = []
        if len(bboxes_In) != 0:
            max_bbox_In, landmarks_In, max_area = getMaxfacebox(bboxes_In)
            if max_area > 8000:          
                landmarks_In = np.array([landmarks_In["left_eye"][0], landmarks_In["right_eye"][0], landmarks_In["nose"][0], landmarks_In["mouth_left"][0], landmarks_In["mouth_right"][0],
                                     landmarks_In["left_eye"][1], landmarks_In["right_eye"][1], landmarks_In["nose"][1], landmarks_In["mouth_left"][1], landmarks_In["mouth_right"][1]])
                landmarks_In = landmarks_In.reshape((2, 5)).T
                nimg = face_preprocess.preprocess(frame_In, max_bbox_In, landmarks_In, image_size='112,112')
                nimg = cv2.cvtColor(nimg, cv2.COLOR_BGR2RGB)
                nimg = np.transpose(nimg, (2, 0, 1))
                embedding_In = embedding_model.get_feature(nimg).reshape(1, -1)

                text_In = "Unknown-"

                # Calculate cosine similarity
                cos_similarity_In, j_In = CosineSimilarity(embedding_In, embeddings)
                print("cos_similarity:" + str(cos_similarity_In))
                if not previous_embedding:
                    previous_cos_similarity_In = cosine_threshold
                    cosine_threshold = 0.8
                    proba_threshold = 0.8
                else:
                    cosine_threshold = 0.9
                    previous_cos_similarity_In, _ = CosineSimilarity(embedding_In, previous_embedding)

                if cos_similarity_In < cosine_threshold:
                    name = labels[j_In]
                    text_In = "{}".format(name)
                    #print("In - when=1")
                    #print(cos_similarity_In)
                else:
                    # print("I don't know you")
                    if previous_cos_similarity_In >= proba_threshold:#cosine_threshold:
                        print("Unknown_new_face")
                        text_label = "Person" + str(indexNumber)
                        text_In += text_label
                        embeddings = np.concatenate((embeddings, embedding_In))
                        labels = np.append(labels, text_label)
                        indexNumber += 1
                        previous_embedding.append(embedding_In)
                        new_labels.append(text_label)
                        new_embeddings.append(embedding_In)
                        #print("In - when=2")
                    else:
                        _, k = CosineSimilarity(embedding_In, previous_embedding)
                        name = new_labels[k]
                        text_In = "{}".format(name)
                        #print("In - when=3")
                    #print(cos_similarity_In)
                    print(previous_cos_similarity_In)

                # Start tracking
                tracker = dlib.correlation_tracker()
                tx0, ty0, tx1, ty1 = (max_bbox_In[0], max_bbox_In[1], max_bbox_In[2], max_bbox_In[3])
                rect = dlib.rectangle(int(tx0), int(ty0), int(tx1), int(ty1))
                tracker.start_track(rgb_In, rect)
                trackers_In.append(tracker)
                texts_In.append(text_In)

                y = max_bbox_In[1] - 10 if max_bbox_In[1] - 10 > 10 else max_bbox_In[1] + 10
                cv2.putText(frame_In, text_In, (max_bbox_In[0], y), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)
                cv2.rectangle(frame_In, (max_bbox_In[0], max_bbox_In[1]), (max_bbox_In[2], max_bbox_In[3]), FACEBOX_COLOR, 2)

        else:
            for tracker, text_In in zip(trackers_In,texts_In):
                pos = tracker.get_position()

                # unpack the position object
                startX = int(pos.left())
                startY = int(pos.top())
                endX = int(pos.right())
                endY = int(pos.bottom())

                cv2.rectangle(frame_In, (startX, startY), (endX, endY), FACEBOX_COLOR, 2)
                cv2.putText(frame_In, text_In, (startX, startY - 15), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0,0,255),2)
                
        trackers = []
        texts = []
        if len(bboxes_Out) != 0:
            #reco_tick = time.time()
            max_bbox, landmarks, max_area = getMaxfacebox(bboxes_Out)
            if max_area > 8000:
                landmarks = np.array([landmarks["left_eye"][0], landmarks["right_eye"][0], landmarks["nose"][0], landmarks["mouth_left"][0], landmarks["mouth_right"][0],
                                     landmarks["left_eye"][1], landmarks["right_eye"][1], landmarks["nose"][1], landmarks["mouth_left"][1], landmarks["mouth_right"][1]])
                landmarks = landmarks.reshape((2, 5)).T
                nimg = face_preprocess.preprocess(frame, max_bbox, landmarks, image_size='112,112')
                nimg = cv2.cvtColor(nimg, cv2.COLOR_BGR2RGB)
                nimg = np.transpose(nimg, (2, 0, 1))
                embedding = embedding_model.get_feature(nimg).reshape(1, -1)

                text = "Unknown-"

                # Calculate cosine similarity
                cos_similarity, j = CosineSimilarity(embedding, embeddings)
                print(cos_similarity)
                if not previous_embedding:
                    previous_cos_similarity = cosine_threshold
                else:
                    previous_cos_similarity, _ = CosineSimilarity(embedding, previous_embedding)

                if cos_similarity < cosine_threshold:
                    name = labels[j]
                    text = "{}".format(name)
                    when = 1
                else:
                    # print("I don't know you")
                    print(previous_cos_similarity)
                    if previous_cos_similarity >= proba_threshold:#cosine_threshold:
                        print("Unknown_new_face at Output camera")
                        name = None
                        when = 2
                    else:
                        _, k = CosineSimilarity(embedding, previous_embedding)
                        name = new_labels[k]
                        text = "{}".format(name)
                        when = 3
                if name is not None:
                    draw_points, points, hand_box = hand_detector(rgb)
                    #hand score
                    handScore = hand_detector.getHandScore(draw_points, points)
                    if handScore != 0:
                        strScore = "Hand Score : " + str(handScore)
                        cnt.append(handScore)
                        cv2.putText(frame, strScore, (int(10), int(25)), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 2)
                    if draw_points is not None:
                        #draw land mark
                        for point in draw_points:
                            x, y = point
                            cv2.circle(frame, (int(x), int(y)), THICKNESS * 2, POINT_COLOR, THICKNESS)
                        #draw connection
                        for connection in connections:
                            x0, y0 = draw_points[connection[0]]
                            x1, y1 = draw_points[connection[1]]
                            cv2.line(frame, (int(x0), int(y0)), (int(x1), int(y1)), CONNECTION_COLOR, THICKNESS)
                        #draw hand box
                        for i in range(0, 4):
                            x0, y0 = hand_box[i]
                            if i != 3:
                                x1, y1 = hand_box[i + 1]
                            else:
                                x1, y1 = hand_box[0]
                            cv2.line(frame, (int(x0), int(y0)), (int(x1), int(y1)), (255, 255, 255), THICKNESS)


                # Start tracking
                tracker = dlib.correlation_tracker()
                tx0, ty0, tx1, ty1 = (max_bbox[0], max_bbox[1], max_bbox[2], max_bbox[3])
                rect = dlib.rectangle(int(tx0), int(ty0), int(tx1), int(ty1))
                # rect = dlib.rectangle(max_bbox[0], max_bbox[1], max_bbox[2], max_bbox[3])
                tracker.start_track(rgb, rect)
                trackers.append(tracker)
                texts.append(text)

                y = max_bbox[1] - 10 if max_bbox[1] - 10 > 10 else max_bbox[1] + 10
                cv2.putText(frame, text, (max_bbox[0], y), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)
                cv2.rectangle(frame, (max_bbox[0], max_bbox[1]), (max_bbox[2], max_bbox[3]), FACEBOX_COLOR, 2)

        else:
            for tracker, text in zip(trackers,texts):
                pos = tracker.get_position()

                # unpack the position object
                startX = int(pos.left())
                startY = int(pos.top())
                endX = int(pos.right())
                endY = int(pos.bottom())

                cv2.rectangle(frame, (startX, startY), (endX, endY), FACEBOX_COLOR, 2)
                cv2.putText(frame, text, (startX, startY - 15), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0,0,255),2)
                if text != "Unknown-" and draw_points is not None:
                    for point in draw_points:
                        x, y = point
                        cv2.circle(frame, (int(x), int(y)), THICKNESS * 2, POINT_COLOR, THICKNESS)
                    for connection in connections:
                        x0, y0 = draw_points[connection[0]]
                        x1, y1 = draw_points[connection[1]]
                        cv2.line(frame, (int(x0), int(y0)), (int(x1), int(y1)), CONNECTION_COLOR, THICKNESS)
                    if handScore != 0:
                        strScore = "Hand Score : " + str(handScore)
                        cv2.putText(frame, strScore, (int(10), int(25)), cv2.FONT_HERSHEY_SIMPLEX, 1, (255,255,255),2)
        print(name)
        cv2.namedWindow("Input_cam")
        cv2.namedWindow("Output_cam")
        cv2.moveWindow("Input_cam", 40, 30)
        cv2.moveWindow("Output_cam", 840, 30)
        frame = cv2.resize(frame, (320, 240))
        frame_In = cv2.resize(frame_In, (320, 240))
        cv2.imshow("Output_cam", frame)
        cv2.imshow("Input_cam", frame_In)
        # cv2.imshow("ipFrame", ipframe)
        # video_out.write(frame)
        # print("Faces detection time: {}s".format(detect_tock-detect_tick))
        # print("Faces recognition time: {}s".format(reco_tock-reco_tick))
        key = cv2.waitKey(1) & 0xFF

        if name is not None and handScore != 0:

            score_number.append(handScore)

            if prev_name != name and prev_name is not None:
                # write_ws.append([1, name, handScore])
                id = len(score_number) - 1
                print("id")
                print(id)
                if id >= 3:
                    score_number2 = [] #200711
                    # find continuous scores
                    for x in range(id, 2, -1): #range(3, id):
                        curr_score = score_number[x-1]
                        pre_score = score_number[x-2]
                        prepre_score = score_number[x-3]
                        if curr_score == pre_score == prepre_score:
                            score_number2.append(curr_score) #200711
                    # pick the most common score
                    print(score_number2)
                    if score_number2 :
                        final_score = most_frequent(score_number2) #200711
                        print(prev_name)
                        print(final_score)
                        write_ws.append([1, prev_name, final_score])
                        for i in range(2, write_ws.max_row):
                            if write_ws.cell(row=i, column=2).value == prev_name and write_ws.cell(row=i, column=4).value != "renew":
                                write_ws.cell(row=i, column=4).value = "renew"
                        
                        # write_ws.cell(row=write_ws.max_row, column=5).value = when

                first_score = score_number[id]
                score_number = []
                score_number.append(first_score)
            prev_name = name
        if key == ord("q"):
            break

#print(name)
#print(score_number)
if prev_name is not None and score_number :  #if name is not None and score_number != 0:
    id = len(score_number) - 1
    #print("id")
    #print(id)
    if id >= 3:
        score_number2 = []  # 200711
        for x in range(id, 2, -1): #range(3, id):
            curr_score = score_number[x-1]
            pre_score = score_number[x-2]
            prepre_score = score_number[x-3]
            if curr_score == pre_score == prepre_score:
                score_number2.append(curr_score)  # 200711
        print(score_number2)
        if score_number2 :
            final_score = most_frequent(score_number2)  # 200711
            print(prev_name)
            print(final_score)
            write_ws.append([1, prev_name, final_score])
            for i in range(2, write_ws.max_row):
                if write_ws.cell(row=i, column=2).value == prev_name and write_ws.cell(row=i, column=4).value != "renew":
                    write_ws.cell(row=i, column=4).value = "renew"
            #write_ws.cell(row=write_ws.max_row, column=5).value = when
        
#video_out.release()
cap.release()
# ipcap.release()
# save to output
print(np.shape(embeddings))
new_data = {"embeddings": embeddings, "names": labels}
f = open(args.embeddings, "wb")
f.write(pickle.dumps(new_data))
f.close()
write_wb.save('excel_test.xlsx')
cv2.destroyAllWindows()
